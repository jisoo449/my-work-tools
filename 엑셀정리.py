import os
import tarfile
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook

# 해당 경로에서 *_secure.tar 파일 찾기
def unzip(dir):
    for file_name in os.listdir(dir):
        if file_name.endswith("_secure.tar.gz"):
            file_path = os.path.join(dir, file_name)
            
            # 압축 해제할 폴더 생성 (파일명과 동일한 폴더)
            extract_dir = os.path.join(dir, file_name.replace(".tar.gz", ""))
            os.makedirs(extract_dir, exist_ok=True)

            # tar 파일 압축 해제
            with tarfile.open(file_path, "r") as tar:
                tar.extractall(path=extract_dir)
                print(f"압축 해제 완료: {file_path} → {extract_dir}")


def writeExcel(dir):
    excel_path = os.path.join(dir, "secure_logs.xlsx")
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 로그 형식을 정규 표현식으로 정의
    log_pattern = re.compile(r"^(\w+ \d+ \d+:\d+:\d+) (\S+) (\S+): (.+)$")

    # 지정된 경로 안의 디렉토리 찾기
    for dir_name in os.listdir(dir):
        dir_path = os.path.join(dir, dir_name)
        
        if os.path.isdir(dir_path):  # 디렉토리인지 확인
            secure_file_path = os.path.join(dir_path, "secure")

            if os.path.exists(secure_file_path):  # secure 파일이 존재하면 처리
                logs = []
                with open(secure_file_path, "r", encoding="utf-8") as f:
                    for line in f:
                        match = log_pattern.match(line.strip())
                        if match:
                            logs.append(match.groups())  # (날짜, 서버명, 권한, 내용) 형태로 저장

                # 데이터프레임으로 변환 후 엑셀 시트에 저장
                if logs:
                    df = pd.DataFrame(logs, columns=["날짜", "서버명", "권한", "내용"])
                    ws = wb.create_sheet(title=dir_name)
                    for row in df.itertuples(index=False, name=None):
                        ws.append(row)

    # 엑셀 저장
    wb.save(excel_path)
    print(f"엑셀 파일 저장 완료: {excel_path}")    


def findUserState(dir):

    # 정규 표현식 패턴 정의
    user_pattern = re.compile(r"session (opened|closed) for user (\S+)")

    # 엑셀 파일 열기
    wb = load_workbook(dir)

    # 모든 시트에 대해 반복
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # E, F 컬럼 제목 추가
        ws["E1"] = "사용자명"
        ws["F1"] = "접속 상태"
        
        for row in range(2, ws.max_row + 1):  # 2번째 행부터 읽기 (1번째 행은 헤더)
            log_entry = ws[f"D{row}"].value  # D열(로그 내용) 읽기
            
            if log_entry:
                match = user_pattern.search(log_entry)
                if match:
                    status, user = match.groups()
                    
                    # 사용자명 추가 (E열)
                    ws[f"E{row}"] = user
                    
                    # 접속 상태 추가 (F열)
                    ws[f"F{row}"] = "접속 생성" if status == "opened" else "접속 해제"

    # 엑셀 저장
    wb.save(dir)
    print(f"엑셀 파일 업데이트 완료: {dir}")

dir=r"C:\Users\INNOGRID\Documents\Amaranth10\[울산항만공사 대표홈페이지] 접속기록 및 정책설정 로그"
dir2=r"C:\Users\INNOGRID\Documents\Amaranth10\secure_logs_수정 by Moon.xlsx"
findUserState(dir2)
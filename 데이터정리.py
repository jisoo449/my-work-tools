import openpyxl
import re
import os
import openpyxl.workbook


# 사용자 입력값 (예시)
start_row = 11  # A열의 시작 행 (필요에 따라 수정)
max_search_rows = 280

string_list = ['hwabul-master01', 'hwabul-master02', 'hwabul-master03', 'hwabul-ingress01', 'hwabul-ingress02', 'hwabul-ingress03', 
                'hwabul-node01', 'hwabul-node02', 'hwabul-node03', 'hwabul-node04', 'hwabul-node05', 'hwabul-node06', 'hwabul-node07', 
                'hwabul-node08', 'hwabul-node09', 'hwabul-node10', 'hwabul-handydb', 'hwabul-cnfdb01', 'hwabul-cnfdb02', 'hwabul-extcnfdb', 
                'hwabul-webhwp', 'hwabul-nas', 'SECLOUDiT-Console', 'SECLOUDiT-LB', 'SECLOUDiT-Logging', 'SECLOUDiT-Registry', 'hwabul-v3', 
                'hwabul-ngs', 'hwabul-petra', 'hwabul-cspm1', 'hwabul-webfilter', 'hwabul-commgt'] # 정렬 순서를 위한 문자열 리스트
pattern = r'■ (.*?) \(' #  r'■ ([^(]+) \('


### 파일 이름 읽어오기
def get_filenames(dir):
    """
    ## 파일 이름 읽어오기

    Args:
        dir: 파일 목록록을 조회할 경로
    """
    file_list = os.listdir(dir)
    return file_list


### 리소스 데이터 읽어오기
def read_data(file_path, a_i_data={}):
    # 워크북 로드 및 시트 선택
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 데이터 추출 로직
    for row in range(start_row, max_search_rows + 1, 8):
        s = sheet[f'A{row}'].value
        if s is not None:
            s = re.search(pattern, s).group(1)
            a_value = s.replace('#','').replace(' ','')
            i_value = (sheet[f'H{row + 4}'].value, sheet[f'H{row + 5}'].value)
            a_i_data[a_value] = i_value
            
    wb.close()

    return a_i_data

### 데이터 저장하기
def save_resource(wb, sheet, data):
    # 새 파일 생성 및 데이터 쓰기
    
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet)

        # 헤더 추가 (선택 사항)
        ws['A1'] = '파드명'
        ws['B1'] = 'CPU'  # I열을 B열에 저장 (필요 시 위치 변경)
        ws['C1'] = 'MEM'  # I열을 B열에 저장 (필요 시 위치 변경)

        # 데이터 쓰기
        for i, s in enumerate(string_list):
            ws[f'A{i+2}'] = s

            if data.get(s):
                ws[f'B{i+2}'] = data.get(s)[0]  # I열 데이터를 B열에 저장
                ws[f'C{i+2}'] = data.get(s)[1]  # I열 데이터를 B열에 저장
            else:
                ws[f'B{i+2}'] = '-'
                ws[f'C{i+2}'] = '-'

        print(f'처리 완료! {sheet} 시트가 생성되었습니다.')
    
    else:
        ws = wb[sheet]

        # 헤더 추가 (선택 사항)
        ws['D1'] = '송신'
        ws['E1'] = '수신'

        for i, s in enumerate(string_list):
            if data.get(s):
                ws[f'D{i+2}'] = data.get(s)[1]
                ws[f'E{i+2}'] = data.get(s)[0]

            else:
                ws[f'B{i+2}'] = '-'
                ws[f'C{i+2}'] = '-'
        print(f'처리 완료! {sheet} 시트에 트래픽 내역이 추가가되었습니다.')


### 여러 월의 리소스 추출
def extract_resources():
    """
    ## 여러 월의 리소스 추출
    리소스(CPU, MEM) 사용률을 읽어와 추출 및 저장하는 함수.  
    특정 디렉토리 안에 있는 여러 달의 정보를 가져오며,  
    새로운 파일을 만들 때 사용한다. 
    """
    # 1. 파일 읽어오기
    resource_file_list = get_filenames("./리소스")

    # 2. 편집할 파일 열기
    filepath_new = "데이터 추출.xlsx"
    wb_new = openpyxl.Workbook()

    # 3. 파일별 데이터 가져오기
    for filename in resource_file_list:   
        data = read_data(f'./리소스/{filename}')

        # 4. 데이터 추가하기기
        ym_pattern = re.compile(r"\b\d{4}\b")
        sheet=re.search(ym_pattern, filename).group()
        save_resource(wb_new, sheet, data)

    # 5. 파일 저장 및 닫기
    wb_new.save(filepath_new)
    wb_new.close()


### 여러 월의 트래픽 추출
def extract_traffics():
    """
    ## 여러 월의 트래픽 추출
    네트워크 송수신량을 읽어와 추출 및 저장하는 함수.  
    특정 디렉토리 안에 있는 여러 달의 정보를 가져오며,  
    **반드시 extract_resources 실행 후에 동작해야 한다.**
    """
    resource_file_list = get_filenames("./네트워크")
    filepath = "데이터 추출.xlsx"
    wb = openpyxl.load_workbook(filepath)

    for filename in resource_file_list:
        if 'part2' not in filename:
            if 'part1' in filename:
                data = read_data(f'./네트워크/{filename}')
                data = read_data(f'./네트워크/{filename.replace('part1', 'part2')}', data)
                
            else: 
                data = read_data(f'./네트워크/{filename}')
            
            print(data)

            # 4. 데이터 추가하기기
            ym_pattern = re.compile(r"\b\d{4}\b")
            sheet=re.search(ym_pattern, filename).group()
            save_resource(wb, sheet, data)
 
    # 5. 파일 저장 및 닫기
    wb.save(filepath)
    wb.close()


### 특정 월 리소스/트래픽 추출
def extract_network_resource(ym, resource_filename, network_filename1, network_filename2=''):
    """
    특정  월의 리소스,트래픽 추출출
    특정 월의 엑셀 파일만 추출하는 함수.  
    파드의 리소스(CPU, MEM) 사용량, 네트워크 송수신 트래픽 정보
    Args:
        ym: 연도-월. 시트명이 된다.  
        resource_filename: 리소스(CPU, MEM) 정보가 저장된 엑셀파일
        network_filename1: 네트워크 송수신 정보가 저장된 엑셀파일
        network_filename2: 네트워크 송수신 정보가 저장된 엑셀파일 part2. 정보가 많아 두 개 파일로 나눌 때 사용되는 인자이다.
    """

    # 1. 파일 생성 및 열기
    dir=r'C:\Users\INNOGRID\Documents\Amaranth10\NIA 월간운영보고서 재작성 요청'
    filepath = os.path.join(dir,"데이터 추출.xlsx")
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.Workbook()

    # 2. 리소스 추출 및 파일 저장
    resource_data = read_data(resource_filename)  
    save_resource(wb, ym, resource_data)
    
    # 3. 네트워크 추출 및 파일 저장
    network_data = read_data(network_filename1)
    if network_filename2 != '':
        network_data = read_data(network_filename2, network_data)
    save_resource(wb, ym, network_data)

    wb.save(filepath)
    wb.close()


dir=r'C:\Users\INNOGRID\Documents\Amaranth10\NIA 월간운영보고서 재작성 요청'
resource="리소스-2412.xlsx"
netework1="화성-네트워크-2412-part1.xlsx"
netework2="화성-네트워크-2412-part2.xlsx"
r_file=os.path.join(dir, resource)
n_file=os.path.join(dir, netework1)
n_file2=os.path.join(dir, netework2)

extract_network_resource('2412', r_file, n_file, n_file2)

#수기로 하면 월 하나 당 20분 걸렸는데, 클릭 한 번으로 되니까 거의 1초면 됨
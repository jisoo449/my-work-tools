import re
import io
from datetime import date
from dateutil.relativedelta import relativedelta
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from xlsx_to_ppt import build_ppt_from_template
from openpyxl import load_workbook
from merge_ppt import merge_ppts_with_merger

from duplicate_ppts import copy_pptx_to_multiple_names

# -----------------------------
# 0) 사용자 환경에 맞게 수정할 설정
# -----------------------------
CPU_MEM_XLSX = "cpu_mem.xlsx"
NETWORK_XLSX = "network.xlsx"
FS_XLSX = "filesystem.xlsx"
REPORT_OUTPUT_PPTX = "server_report.pptx"
TEMPLATE_PPTX = "template02.pptx"

# 서버 블록을 찾을 때 사용할 패턴 (예: bastion-lnx (172.25.0.74))
SERVER_NAME_PATTERN = re.compile(r".+\(\s*\d{1,3}(?:\.\d{1,3}){3}\s*\)")

# 서버 블록 범위(시작행부터 몇 행까지를 같은 서버 섹션으로 볼지)
# Excel 레이아웃에 따라 필요 시 늘리세요.
DEFAULT_BLOCK_HEIGHT = 60

# Min/Max/Avg 라벨 후보 (영/한 혼용 대응)
LABEL_MIN = {"min", "minimum", "최소"}
LABEL_MAX = {"max", "maximum", "최대"}
LABEL_AVG = {"avg", "average", "평균", "mean"}

# -----------------------------
# 1) 데이터 구조
# -----------------------------
@dataclass
class Stats:
    min: Optional[float] = None
    max: Optional[float] = None
    avg: Optional[float] = None

@dataclass
class MetricBlock:
    # CPU/MEM 은 2개 지표, Network는 IN/OUT 2개 지표 등 복수 가능
    stats_by_metric: Dict[str, Stats]     # e.g. {"CPU": Stats(...), "MEM": Stats(...)}
    images: List[bytes]                  # 이미지(PNG/JPG 바이너리). 블록 내 이미지 순서대로 담김

@dataclass
class ServerReport:
    server_name: str
    cpu_mem: MetricBlock
    network: MetricBlock
    filesystem: MetricBlock

# -----------------------------
# 2) 원하는 column 삭제
# -----------------------------
def delete_columns(files, columns_to_delete):
    for file in files:
        wb = load_workbook(file)
        ws = wb.active
        for col in sorted(columns_to_delete, reverse=True):
            ws.delete_cols(col)
        wb.save(file)

# -----------------------------
# 2) Excel 파싱 유틸
# -----------------------------
def _normalize_label(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()

def _is_server_header(value) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    return bool(SERVER_NAME_PATTERN.fullmatch(s))

def _scan_server_headers(ws, header_col=1) -> List[Tuple[str, int]]:
    """
    A열(기본)에서 서버 헤더 행을 찾고 [(서버명, 행번호)] 리스트 반환.
    """
    found = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=header_col).value
        if _is_server_header(v):
            found.append((str(v).strip(), row))
    return found

def _block_end(start_row: int, next_start_row: Optional[int], default_height: int) -> int:
    if next_start_row is not None:
        return next_start_row - 1
    return start_row + default_height

def _extract_images_in_row_range(ws, row_start: int, row_end: int) -> List[bytes]:
    """
    openpyxl로 추출 가능한 삽입 이미지들을 row_start~row_end 범위에서 찾아 bytes로 반환.
    """
    images_bytes: List[bytes] = []
    imgs = getattr(ws, "_images", []) or []
    for img in imgs:
        try:
            anchor = img.anchor
            # OneCellAnchor / TwoCellAnchor 모두 _from 을 가짐
            r = anchor._from.row + 1  # 0-based -> 1-based
        except Exception:
            continue

        if row_start <= r <= row_end:
            bio = io.BytesIO()
            # openpyxl Image는 PIL 이미지 또는 원본을 internal로 가짐.
            # _data()가 있으면 가장 안전.
            if hasattr(img, "_data") and callable(img._data):
                images_bytes.append(img._data())
            else:
                # fallback: PIL로 저장 시도
                try:
                    img.image.save(bio, format="PNG")
                    images_bytes.append(bio.getvalue())
                except Exception:
                    pass
    return images_bytes

def _try_parse_stats_table(ws, row_start: int, row_end: int) -> Dict[str, Stats]:
    """
    블록 내에서 '지표명 / Min / Max / Avg' 형태를 휴리스틱으로 파싱.
    - 지표명은 좌측(예: B열)이나 첫 번째 컬럼에 있을 수 있어, 행 단위로 탐색합니다.
    - 라벨 텍스트(최소/최대/평균 or Min/Max/Avg)가 있는 행/열을 기준으로 인접 셀 값을 가져옵니다.
    Excel 템플릿이 다르면 이 부분만 조정하면 됩니다.
    """
    stats_by_metric: Dict[str, Stats] = {}

    # 1) "Min/Max/Avg" 헤더가 있는 행을 찾는다.
    header_row = None
    header_cols = {}  # {"min": col, "max": col, "avg": col}
    for r in range(row_start, min(row_end, ws.max_row) + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 30) + 1)]
        norm = [_normalize_label(v) for v in row_vals]

        def find_col(candidates: set) -> Optional[int]:
            for idx, val in enumerate(norm, start=1):
                if val in candidates:
                    return idx
            return None

        cmin = find_col(LABEL_MIN)
        cmax = find_col(LABEL_MAX)
        cavg = find_col(LABEL_AVG)

        # 최소한 2개 이상 잡히면 "헤더 행"으로 간주
        hit = sum(x is not None for x in [cmin, cmax, cavg])
        if hit >= 2:
            header_row = r
            if cmin: header_cols["min"] = cmin
            if cmax: header_cols["max"] = cmax
            if cavg: header_cols["avg"] = cavg
            break

    if header_row is None:
        return stats_by_metric  # 못 찾으면 빈 dict 반환

    # 2) 헤더 바로 아래 몇 줄을 지표 행으로 파싱 (빈 줄/다음 섹션 만나면 중단)
    for r in range(header_row + 1, min(header_row + 15, row_end) + 1):
        metric_name = ws.cell(row=r, column=1).value  # 기본: A열에 지표명
        if metric_name is None:
            # A열이 비었는데 B열에 지표명이 있을 수 있음
            metric_name = ws.cell(row=r, column=2).value

        if metric_name is None:
            continue

        mname = str(metric_name).strip()
        # 서버 헤더가 다시 나오면 중단
        if _is_server_header(mname):
            break

        # 숫자값 추출
        st = Stats()
        for k, c in header_cols.items():
            val = ws.cell(row=r, column=c).value
            if val is None:
                val = None
            setattr(st, k, val)

        # "의미있는 값"이 하나라도 있으면 등록
        if any(getattr(st, k) is not None for k in ["min", "max", "avg"]):
            # 지표명 표준화(필요시 사용자가 여기서 매핑)
            stats_by_metric[mname] = st

    return stats_by_metric

def parse_excel_as_blocks(xlsx_path: str, sheet_name: Optional[str] = None,
                          default_height: int = DEFAULT_BLOCK_HEIGHT) -> Dict[str, MetricBlock]:
    """
    Excel 한 파일을 서버명 기준 MetricBlock으로 파싱하여 dict 반환.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = _scan_server_headers(ws, header_col=1)
    if not headers:
        raise RuntimeError(f"[{xlsx_path}] A열에서 서버 헤더를 찾지 못했습니다. (패턴/열 위치 확인 필요)")

    results: Dict[str, MetricBlock] = {}
    for i, (server, start_row) in enumerate(headers):
        next_row = headers[i + 1][1] if i + 1 < len(headers) else None
        end_row = _block_end(start_row, next_row, default_height)

        images = _extract_images_in_row_range(ws, start_row, end_row)
        stats = _try_parse_stats_table(ws, start_row, end_row)

        server_trimed = re.sub(r'^\s*■\s*|\s*\([^)]*\)\s*$', '', server).strip()

        results[server_trimed] = MetricBlock(stats_by_metric=stats, images=images)

    return results


# -----------------------------
# 4) 메인 실행부
# -----------------------------
def main():

    date_str = (date.today() - relativedelta(months=1)).strftime("%y년 %#m월")
    OUTPUT_PPTX = f"{date_str} 월간 운영보고서.pptx"

    cpu_mem_blocks = parse_excel_as_blocks(CPU_MEM_XLSX)
    net_blocks = parse_excel_as_blocks(NETWORK_XLSX)
    fs_blocks = parse_excel_as_blocks(FS_XLSX)

    # 서버 키(서버명 (IP)) 기준으로 교집합/합집합 구성
    all_servers = sorted(set(cpu_mem_blocks) | set(net_blocks) | set(fs_blocks))
    
    reports: List[ServerReport] = []
    for s in all_servers:
        reports.append(
            ServerReport(
                server_name=s,
                cpu_mem=cpu_mem_blocks.get(s, MetricBlock({}, [])),
                network=net_blocks.get(s, MetricBlock({}, [])),
                filesystem=fs_blocks.get(s, MetricBlock({}, [])),
            )
        )

    build_ppt_from_template(
        template_pptx=TEMPLATE_PPTX,
        output_pptx=REPORT_OUTPUT_PPTX,
        reports=reports,
        date=date_str
    )

    merge_ppts_with_merger(["template01.pptx", REPORT_OUTPUT_PPTX,"template03.pptx"],OUTPUT_PPTX)
    
    agencies = ["경북농식품유통교육진흥원","경북문화재단","경북바이오산업연구원","경북여성정책개발원","경북종합자원봉사센터","경북행복재단","경상북도경제진흥원","경상북도교통문화연수원","경상북도인재평생교육재단","경상북도장애인체육회","경상북도호국보훈재단","경상북도환경연수원","독도재단","새마을재단","한국국학진흥원"]
    for agency in agencies:
        copy_pptx_to_multiple_names(
            date=date_str,
            src_pptx=OUTPUT_PPTX,
            target_pptx=agency)

if __name__ == "__main__":
    main()